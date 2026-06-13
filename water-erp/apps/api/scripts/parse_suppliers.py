#!/usr/bin/env python3
"""
Parse the two supplier Excel files into a single normalized JSON for DB import.

- 供应商入库表（总表）.xlsx: rich master list (creditCode, legalPerson, address, ...)
- 供应商入库表（生技部）.xlsx: qualified vendor directory (cert level + recommended scope only)

Output: suppliers_import.json — one object per unique supplier, merged by name.
"""
import json
import re
import openpyxl

KNOWLEDGE = "/Users/qihao/Desktop/ERP/water-erp/apps/api/knowledge"
OUT = "/Users/qihao/Desktop/ERP/water-erp/apps/api/scripts/suppliers_import.json"


def norm_name(s):
    """Match the app's normalizedName rule (trim().toLowerCase()) + punctuation fold."""
    if s is None:
        return ""
    return (str(s).strip().lower()
            .replace("（", "(").replace("）", ")")
            .replace(" ", "").replace("　", ""))


def clean_text(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def split_multi(s):
    """Split a cell that may hold multiple values separated by ；;，, or newline."""
    if s is None:
        return []
    parts = re.split(r"[；;，,、\n]+", str(s))
    return [p.strip() for p in parts if p and p.strip()]


def clean_phone(s):
    if s is None:
        return ""
    # keep digits, strip anything else; Excel may store as float
    if isinstance(s, float):
        s = str(int(s))
    return re.sub(r"[^\d\-+]", "", str(s)).strip()


def read_master():
    """总表 → dict[normalizedName] = {fields}"""
    wb = openpyxl.load_workbook(f"{KNOWLEDGE}/供应商入库表（总表）.xlsx", data_only=True)
    ws = wb.active
    hdr = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(hdr)}
    out = {}
    dups = 0
    for r in range(3, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        name = clean_text(v[idx["单位名称"]])
        if not name:
            continue
        key = norm_name(name)
        if key in out:
            dups += 1
            continue  # keep first occurrence
        contacts = []
        cnames = split_multi(v[idx["联系人"]])
        cphones = split_multi(v[idx["联系电话"]])
        for i, cn in enumerate(cnames):
            ph = clean_phone(cphones[i]) if i < len(cphones) else ""
            contacts.append({"name": cn, "phone": ph})
        out[key] = {
            "name": name,
            "creditCode": clean_text(v[idx["社会统一信用代码"]]),
            "enterpriseType": clean_text(v[idx["单位性质"]]),
            "legalPerson": clean_text(v[idx["法定代表人"]]),
            "registeredAddress": clean_text(v[idx["单位地址"]]),
            "businessType": clean_text(v[idx["业务类型"]]),
            "qualificationText": clean_text(v[idx["资质/许可/认证类信息"]]),
            "remark": clean_text(v[idx["备注"]]),
            "riskHint": clean_text(v[idx["审计法务部风险提示"]]),
            "contacts": contacts,
        }
    print(f"总表: {len(out)} unique suppliers ({dups} duplicate names skipped)")
    return out


def read_directory():
    """生技部 → dict[normalizedName] = {fields}"""
    wb = openpyxl.load_workbook(f"{KNOWLEDGE}/供应商入库表（生技部）.xlsx", data_only=True)
    ws = wb.active
    hdr = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(hdr)}
    out = {}
    dups = 0
    for r in range(3, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        name = clean_text(v[idx["单位名称"]])
        if not name:
            continue
        key = norm_name(name)
        if key in out:
            dups += 1
            continue
        out[key] = {
            "name": name,
            "certLevel": clean_text(v[idx["持证等级"]]),
            "recommendedScope": clean_text(v[idx["推荐业务范围"]]),
            "remark": clean_text(v[idx["备注"]]),
            "formerName": clean_text(v[idx["原名"]]),
        }
    print(f"生技部: {len(out)} unique suppliers ({dups} duplicate names skipped)")
    return out


def merge(master, directory):
    merged = []
    both = 0
    dir_only = 0
    # master first (full data)
    for key, m in master.items():
        d = directory.get(key)
        sources = ["总表"]
        scope_parts = []
        if m["businessType"]:
            scope_parts.append(m["businessType"])
        qual_parts = []
        if m["qualificationText"]:
            qual_parts.append("资质/许可/认证: " + m["qualificationText"])
        if d:
            sources.append("生技部")
            both += 1
            if d["recommendedScope"]:
                scope_parts.append("推荐业务范围: " + d["recommendedScope"])
            if d["certLevel"]:
                qual_parts.append("持证等级: " + d["certLevel"])
        merged.append({
            "name": m["name"],
            "creditCode": m["creditCode"] or None,
            "enterpriseType": m["enterpriseType"],
            "legalPerson": m["legalPerson"],
            "registeredAddress": m["registeredAddress"],
            "businessScope": "；".join(scope_parts) if scope_parts else "",
            "businessType": m["businessType"],
            "qualificationText": "\n".join(qual_parts) if qual_parts else None,
            "contacts": m["contacts"],
            "remark": m["remark"],
            "riskHint": m["riskHint"] or None,
            "sources": sources,
        })
    # directory-only
    for key, d in directory.items():
        if key in master:
            continue
        dir_only += 1
        merged.append({
            "name": d["name"],
            "creditCode": None,
            "enterpriseType": "",
            "legalPerson": "",
            "registeredAddress": "",
            "businessScope": ("推荐业务范围: " + d["recommendedScope"]) if d["recommendedScope"] else "",
            "businessType": None,
            "qualificationText": ("持证等级: " + d["certLevel"]) if d["certLevel"] else None,
            "contacts": [],
            "remark": d["remark"],
            "riskHint": None,
            "sources": ["生技部"],
        })
    print(f"Merge: {both} in both, {dir_only} directory-only → {len(merged)} total")
    return merged


def main():
    master = read_master()
    directory = read_directory()
    merged = merge(master, directory)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    # sanity stats
    with_cc = sum(1 for s in merged if s["creditCode"])
    with_contact = sum(1 for s in merged if s["contacts"])
    with_qual = sum(1 for s in merged if s["qualificationText"])
    print(f"\nWrote {len(merged)} suppliers → {OUT}")
    print(f"  with creditCode : {with_cc}")
    print(f"  with contacts   : {with_contact}")
    print(f"  with qual text  : {with_qual}")


if __name__ == "__main__":
    main()
